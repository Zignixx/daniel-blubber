"""
CiteCut
=======

Streamlit-Oberfläche: Word/PDF hochladen, Kapitel auswählen, Wörter zählen –
mit Ausschluss von Klammer-Zitaten und Tabellen sowie Optionen für
Überschriften und Fußnoten.

Start:
    streamlit run blubberzaehler.py
"""

import csv
import io
import json
import time

import streamlit as st
import streamlit.components.v1 as components

from blubber_core import (
    CountNode,
    CountOptions,
    count_document,
    grand_total,
    parse_document,
)


st.set_page_config(page_title="CiteCut", page_icon="✂️", layout="centered")


# --------------------------------------------------------------------------- #
# Kapitel-Übersicht: funnel-artige Balkenleiste + aufklappbare Baumansicht
# --------------------------------------------------------------------------- #
_TREE_HTML = """
<style>
  :root {
    --bz-track: #eceae3;
    --bz-text: #0b0b0b;
    --bz-text-secondary: #52514e;
    --bz-text-muted: #898781;
    --bz-surface: #fcfcfb;
    --bz-hover: rgba(11,11,11,0.06);
    --bz-guide: rgba(11,11,11,0.20);
    --bz-c0: #2a78d6; --bz-c1: #eb6834; --bz-c2: #1baf7a; --bz-c3: #eda100;
    --bz-c4: #e87ba4; --bz-c5: #008300; --bz-c6: #4a3aa7; --bz-c7: #e34948;
  }
  @media (prefers-color-scheme: dark) {
    :root {
      --bz-track: #2c2c2a;
      --bz-text: #ffffff;
      --bz-text-secondary: #c3c2b7;
      --bz-text-muted: #898781;
      --bz-surface: #1a1a19;
      --bz-hover: rgba(255,255,255,0.09);
      --bz-guide: rgba(255,255,255,0.20);
      --bz-c0: #3987e5; --bz-c1: #d95926; --bz-c2: #199e70; --bz-c3: #c98500;
      --bz-c4: #d55181; --bz-c5: #008300; --bz-c6: #9085e9; --bz-c7: #e66767;
    }
  }
  html, body { margin: 0; background: transparent; }
  .bz-wrap {
    font-family: -apple-system, "Segoe UI", system-ui, sans-serif;
    color: var(--bz-text);
  }
  .bz-c0 { background: var(--bz-c0); } .bz-c1 { background: var(--bz-c1); }
  .bz-c2 { background: var(--bz-c2); } .bz-c3 { background: var(--bz-c3); }
  .bz-c4 { background: var(--bz-c4); } .bz-c5 { background: var(--bz-c5); }
  .bz-c6 { background: var(--bz-c6); } .bz-c7 { background: var(--bz-c7); }

  .bz-stackbar {
    position: relative; height: __BAR_HEIGHT__px; overflow: visible;
  }
  .bz-stackbar svg { display: block; width: 100%; height: 100%; }
  .bz-stackbar path { transition: filter .12s; }
  .bz-seg-hit {
    position: absolute; top: 0; height: 100%; cursor: default;
  }
  .bz-seg-hit::after {
    content: attr(data-tip);
    position: absolute; top: calc(100% + 8px); left: 50%;
    transform: translateX(-50%) translateY(-4px);
    background: var(--bz-text); color: var(--bz-surface);
    font-size: 12px; padding: 6px 10px; border-radius: 6px;
    white-space: nowrap; opacity: 0; pointer-events: none;
    transition: opacity .12s, transform .12s; z-index: 10;
  }
  .bz-seg-hit:hover::after { opacity: 1; transform: translateX(-50%) translateY(0); }

  .bz-toolbar { display: flex; justify-content: flex-end; gap: 14px; margin: 8px 2px; }
  .bz-link {
    font-size: 12px; color: var(--bz-text-secondary); background: none; border: none;
    cursor: pointer; padding: 0; font-family: inherit;
  }
  .bz-link:hover { color: var(--bz-text); text-decoration: underline; }

  .bz-col-headers {
    display: flex; align-items: center; gap: 8px;
    padding: 0 6px 4px; margin-top: 4px;
  }
  .bz-col-headers-spacer { flex: 1 1 auto; }
  .bz-col-headers-meta {
    display: flex; align-items: center; gap: 10px; flex: none;
    font-size: 10.5px; font-weight: 600; text-transform: uppercase; letter-spacing: .04em;
    color: var(--bz-text-muted);
  }
  .bz-col-header-anteil { width: 120px; }
  .bz-col-header-woerter { text-align: right; }

  .bz-tree { position: relative; overflow: visible; }
  .bz-hover-highlight {
    position: absolute; top: 0; left: 0; width: 0; height: 0;
    border-radius: 6px; background: var(--bz-hover);
    opacity: 0; pointer-events: none; z-index: 5;
    transition: opacity .12s ease,
      top .16s cubic-bezier(.2,.8,.2,1), left .16s cubic-bezier(.2,.8,.2,1),
      width .16s cubic-bezier(.2,.8,.2,1), height .16s cubic-bezier(.2,.8,.2,1);
  }
  .bz-hover-highlight.bz-hl-visible { opacity: 1; }
  .bz-row {
    position: relative; display: flex; align-items: center; gap: 8px;
    padding: 7px 6px; border-radius: 6px; min-height: 20px;
  }
  .bz-row.bz-has-children { cursor: pointer; }

  /* Baum-Verbindungslinien: eine SVG-Ebene verbindet die Knoten-Kreise direkt
     Kreis-zu-Kreis (Mittelpunkt zu Mittelpunkt), analog zur "animated files"-
     Optik. Die Hauptkapitel-Linie läuft durch alle Hauptkapitel-Kreise und
     verzweigt an jedem Kreis in dessen Unterkapitel. */
  .bz-connectors { position: absolute; top: 0; left: 0; pointer-events: none; z-index: 1; overflow: visible; }
  .bz-edge { fill: none; stroke: var(--bz-guide); stroke-width: 1.75px; }

  .bz-node {
    position: relative; z-index: 2; flex: none;
    width: 14px; height: 14px;
    display: flex; align-items: center; justify-content: center;
  }
  .bz-node.bz-expandable { border-radius: 50%; }
  .bz-node-dot { width: 8px; height: 8px; border-radius: 50%; flex: none; }
  .bz-node svg { transition: transform .18s cubic-bezier(.2,.8,.2,1); }
  .bz-row.bz-open > .bz-node svg { transform: rotate(90deg); }
  .bz-title {
    flex: 1 1 auto; min-width: 0; overflow: hidden; text-overflow: ellipsis; white-space: nowrap;
    font-size: 13.5px;
  }
  .bz-meta { display: flex; align-items: center; gap: 10px; flex: none; }
  .bz-words { font-variant-numeric: tabular-nums; font-size: 12.5px; color: var(--bz-text-secondary); }
  .bz-words-strong { font-weight: 700; color: var(--bz-text); }
  .bz-minibar { width: 64px; height: 6px; border-radius: 3px; background: var(--bz-track); overflow: hidden; }
  .bz-minibar-fill { display: block; height: 100%; border-radius: 3px; }
  .bz-pct {
    font-variant-numeric: tabular-nums; font-size: 12.5px; color: var(--bz-text-secondary);
    width: 46px; text-align: right;
  }

  .bz-children { overflow: hidden; transition: height .2s cubic-bezier(.2,.8,.2,1); }
  .bz-own-row { opacity: .65; font-style: italic; }
  .bz-own-row .bz-title { font-size: 12.5px; }
</style>
<div class="bz-wrap">
  <div class="bz-stackbar" id="bz-bar"></div>
  <div class="bz-toolbar">
    <button type="button" class="bz-link" id="bz-expand-all">Alle aufklappen</button>
    <button type="button" class="bz-link" id="bz-collapse-all">Alle zuklappen</button>
  </div>
  <div class="bz-col-headers">
    <span class="bz-col-headers-spacer"></span>
    <span class="bz-col-headers-meta">
      <span class="bz-col-header-anteil">Anteil</span>
      <span class="bz-col-header-woerter">Wörter</span>
    </span>
  </div>
  <div class="bz-tree" id="bz-tree">
    <svg class="bz-connectors" id="bz-connectors"></svg>
    <div class="bz-hover-highlight" id="bz-hl"></div>
  </div>
</div>
<script>
(function () {
  var DATA = __DATA__;
  var COLOR_COUNT = 8;

  function fmtN(n) { return n.toLocaleString('de-DE'); }
  function fmtPct(p) {
    return p.toLocaleString('de-DE', {minimumFractionDigits: 1, maximumFractionDigits: 1}) + ' %';
  }
  function labelOf(node, fallback) {
    return (node.number ? node.number + ' ' : '') + (node.title || fallback);
  }

  // --- Funnel-artige Balkenleiste (ein Segment pro Hauptkapitel, Wellenform, --- //
  // --- Reihenfolge bleibt die Dokumentreihenfolge, keine Sortierung)       --- //
  var bar = document.getElementById('bz-bar');
  var n = DATA.length;
  var VW = 1000, VH = 34; // virtuerer Koordinatenraum, wird per SVG auf 100% gestreckt
  var gap = 3;
  var segW = (VW - gap * (n - 1)) / n;

  var svgNS = 'http://www.w3.org/2000/svg';
  var svg = document.createElementNS(svgNS, 'svg');
  svg.setAttribute('viewBox', '0 0 ' + VW + ' ' + VH);
  svg.setAttribute('preserveAspectRatio', 'none');
  bar.appendChild(svg);

  function segPath(normStart, normEnd, x, w) {
    var my = VH / 2;
    var h0 = Math.max(normStart, 0.04) * VH * 0.46;
    var h1 = Math.max(normEnd, 0.04) * VH * 0.46;
    var cx = w * 0.55;
    var top = 'M ' + x + ' ' + (my - h0) +
      ' C ' + (x + cx) + ' ' + (my - h0) + ', ' + (x + w - cx) + ' ' + (my - h1) + ', ' + (x + w) + ' ' + (my - h1);
    var bot = 'L ' + (x + w) + ' ' + (my + h1) +
      ' C ' + (x + w - cx) + ' ' + (my + h1) + ', ' + (x + cx) + ' ' + (my + h0) + ', ' + x + ' ' + (my + h0);
    return top + ' ' + bot + ' Z';
  }

  // Auf das größte Kapitel der Auswahl normieren (nicht auf 100%), damit die
  // Ausschläge zwischen den Kapiteln deutlicher sichtbar werden.
  var maxPct = Math.max.apply(null, DATA.map(function (d) { return d.pct; })) || 1;

  var xCursor = 0;
  DATA.forEach(function (node, i) {
    var norm = Math.max(node.pct, 0) / maxPct;
    var normNext = i + 1 < n ? Math.max(DATA[i + 1].pct, 0) / maxPct : norm;

    var path = document.createElementNS(svgNS, 'path');
    path.setAttribute('d', segPath(norm, normNext, xCursor, segW));
    path.style.fill = 'var(--bz-c' + (i % COLOR_COUNT) + ')';
    svg.appendChild(path);

    var hit = document.createElement('div');
    hit.className = 'bz-seg-hit';
    hit.style.left = (xCursor / VW * 100) + '%';
    hit.style.width = (segW / VW * 100) + '%';
    hit.setAttribute('data-tip', labelOf(node, '') + ' · ' + fmtN(node.total) + ' Wörter · ' + fmtPct(node.pct));
    hit.addEventListener('mouseenter', function () { path.style.filter = 'brightness(1.12)'; });
    hit.addEventListener('mouseleave', function () { path.style.filter = ''; });
    bar.appendChild(hit);

    xCursor += segW + gap;
  });

  // --- Baumansicht ---------------------------------------------------------- //
  var treeRoot = document.getElementById('bz-tree');
  var highlight = document.getElementById('bz-hl');
  var connectorsSvg = document.getElementById('bz-connectors');
  var NODE_COL = 20; // Einrückung pro Ebene
  var edges = []; // { from: Kreis-Element, to: Kreis-Element }

  // Wandernde Hover-Hervorhebung (analog zur "FilesHighlight" des animierten
  // Datei-Baum-Primitives): eine einzelne Box gleitet zur jeweils gehoverten
  // Zeile, statt dass jede Zeile einzeln ihren Hintergrund umschaltet. Sie
  // beginnt am Knoten-Kreis der Zeile (nicht am linken Rand des Baums), damit
  // tiefer eingerückte Unterkapitel eine entsprechend kürzere, eingerückte
  // Markierung bekommen statt der vollen Zeilenbreite.
  function showHighlight(row) {
    var treeRect = treeRoot.getBoundingClientRect();
    var rowRect = row.getBoundingClientRect();
    var circle = row.querySelector('.bz-node');
    var circleRect = circle ? circle.getBoundingClientRect() : rowRect;
    var left = circleRect.left - 6;
    highlight.style.top = (rowRect.top - treeRect.top) + 'px';
    highlight.style.left = (left - treeRect.left) + 'px';
    highlight.style.width = (rowRect.right - left) + 'px';
    highlight.style.height = rowRect.height + 'px';
    highlight.classList.add('bz-hl-visible');
  }
  treeRoot.addEventListener('mouseleave', function () {
    highlight.classList.remove('bz-hl-visible');
  });

  function buildMeta(words, pct, colorIdx) {
    var meta = document.createElement('span');
    meta.className = 'bz-meta';

    var pctEl = document.createElement('span');
    pctEl.className = 'bz-pct';
    pctEl.textContent = fmtPct(pct);
    meta.appendChild(pctEl);

    var mini = document.createElement('span');
    mini.className = 'bz-minibar';
    var fill = document.createElement('span');
    fill.className = 'bz-minibar-fill bz-c' + colorIdx;
    fill.style.width = Math.min(Math.max(pct, 0), 100) + '%';
    mini.appendChild(fill);
    meta.appendChild(mini);

    var wordsEl = document.createElement('span');
    wordsEl.className = 'bz-words bz-words-strong';
    wordsEl.textContent = fmtN(words);
    meta.appendChild(wordsEl);

    return meta;
  }

  function buildWordsOnlyMeta(words) {
    var meta = document.createElement('span');
    meta.className = 'bz-meta';
    var wordsEl = document.createElement('span');
    wordsEl.className = 'bz-words';
    wordsEl.textContent = fmtN(words);
    meta.appendChild(wordsEl);
    return meta;
  }

  // Jede Zeile bekommt denselben Kreis-Knoten: aufklappbare Kapitel zeigen
  // einen gefüllten, drehbaren Pfeil; Blätter nur einen kleinen Punkt. Beide
  // sitzen in derselben 14px-Spur, damit die Mittelpunkte exakt fluchten.
  function buildNode(colorIdx, expandable) {
    var node = document.createElement('span');
    node.className = 'bz-node' + (expandable ? ' bz-expandable' : '');
    if (expandable) {
      node.style.background = 'var(--bz-c' + colorIdx + ')';
      node.innerHTML = '<svg width="7" height="7" viewBox="0 0 16 16"><path d="M4.5 3 L11.5 8 L4.5 13 Z" fill="#fff"/></svg>';
    } else {
      var dot = document.createElement('span');
      dot.className = 'bz-node-dot bz-c' + colorIdx;
      node.appendChild(dot);
    }
    return node;
  }

  function buildOwnRow(node, depth, colorIdx) {
    var row = document.createElement('div');
    row.className = 'bz-row bz-own-row';
    row.style.paddingLeft = (depth * NODE_COL + 6) + 'px';

    var circle = buildNode(colorIdx, false);
    row.appendChild(circle);

    var title = document.createElement('span');
    title.className = 'bz-title';
    title.textContent = 'eigener Text ohne Unterkapitel';
    row.appendChild(title);

    row.appendChild(buildWordsOnlyMeta(node.own));
    row.addEventListener('mouseenter', function () { showHighlight(row); });

    return { wrapper: row, circle: circle };
  }

  // Animiertes Auf-/Zuklappen (Höhen-Übergang statt hartem display:none/block).
  function setChildrenOpen(childrenBox, open) {
    if (open) {
      childrenBox.style.height = childrenBox.scrollHeight + 'px';
      var onEnd = function (e) {
        if (e.propertyName !== 'height') return;
        childrenBox.style.height = 'auto';
        childrenBox.removeEventListener('transitionend', onEnd);
      };
      childrenBox.addEventListener('transitionend', onEnd);
    } else {
      childrenBox.style.height = childrenBox.scrollHeight + 'px';
      childrenBox.getBoundingClientRect(); // Reflow erzwingen, damit der Übergang startet
      childrenBox.style.height = '0px';
    }
  }

  function buildRow(node, depth, colorIdx, isMain) {
    var wrapper = document.createElement('div');

    var row = document.createElement('div');
    row.className = 'bz-row';
    row.style.paddingLeft = (depth * NODE_COL + 6) + 'px';
    row.addEventListener('mouseenter', function () { showHighlight(row); });

    var hasChildren = node.children && node.children.length > 0;
    var hasOwnLine = isMain && hasChildren && node.own > 0;

    var circle = buildNode(colorIdx, hasChildren);
    row.appendChild(circle);

    var title = document.createElement('span');
    title.className = 'bz-title';
    var labelText = labelOf(node, '(ohne Titel)');
    title.textContent = labelText;
    title.title = labelText;
    row.appendChild(title);

    row.appendChild(isMain ? buildMeta(node.total, node.pct, colorIdx) : buildWordsOnlyMeta(node.total));
    wrapper.appendChild(row);

    if (hasChildren) {
      var childrenBox = document.createElement('div');
      childrenBox.className = 'bz-children'; // Standardmäßig aufgeklappt.
      var childCircles = [];

      if (hasOwnLine) {
        var ownResult = buildOwnRow(node, depth + 1, colorIdx);
        childrenBox.appendChild(ownResult.wrapper);
        childCircles.push(ownResult.circle);
      }
      node.children.forEach(function (child) {
        var childResult = buildRow(child, depth + 1, colorIdx, false);
        childrenBox.appendChild(childResult.wrapper);
        childCircles.push(childResult.circle);
      });
      wrapper.appendChild(childrenBox);

      // Verzweigung vom Eltern-Kreis zum ersten Kind, danach Kreis-zu-Kreis
      // die Geschwister entlang.
      edges.push({ from: circle, to: childCircles[0] });
      for (var k = 0; k < childCircles.length - 1; k++) {
        edges.push({ from: childCircles[k], to: childCircles[k + 1] });
      }

      row.classList.add('bz-has-children', 'bz-open');
      row.addEventListener('click', function () {
        var opening = !row.classList.contains('bz-open');
        row.classList.toggle('bz-open');
        setChildrenOpen(childrenBox, opening);
        animateConnectors();
      });
    }

    return { wrapper: wrapper, circle: circle };
  }

  var mainCircles = [];
  DATA.forEach(function (node, i) {
    var result = buildRow(node, 0, i % COLOR_COUNT, true);
    treeRoot.appendChild(result.wrapper);
    mainCircles.push(result.circle);
  });
  // Die "Hauptlinie": verbindet alle Hauptkapitel-Kreise nacheinander.
  for (var m = 0; m < mainCircles.length - 1; m++) {
    edges.push({ from: mainCircles[m], to: mainCircles[m + 1] });
  }

  // Liegt ein Knoten innerhalb eines gerade eingeklappten Astes? Dann wird
  // seine Verbindungslinie nicht gezeichnet.
  function isCollapsed(el) {
    var node = el.parentElement;
    while (node && node !== treeRoot) {
      if (node.classList.contains('bz-children') && node.style.height === '0px') {
        return true;
      }
      node = node.parentElement;
    }
    return false;
  }

  function renderConnectors() {
    var treeRect = treeRoot.getBoundingClientRect();
    connectorsSvg.setAttribute('width', treeRect.width);
    connectorsSvg.setAttribute('height', treeRoot.scrollHeight);
    while (connectorsSvg.firstChild) connectorsSvg.removeChild(connectorsSvg.firstChild);

    edges.forEach(function (edge) {
      if (isCollapsed(edge.from) || isCollapsed(edge.to)) return;
      var r1 = edge.from.getBoundingClientRect();
      var r2 = edge.to.getBoundingClientRect();
      var x1 = r1.left + r1.width / 2 - treeRect.left;
      var y1 = r1.top + r1.height / 2 - treeRect.top;
      var x2 = r2.left + r2.width / 2 - treeRect.left;
      var y2 = r2.top + r2.height / 2 - treeRect.top;
      var midY = (y1 + y2) / 2;
      var d = (Math.abs(x1 - x2) < 0.5)
        ? ('M ' + x1 + ' ' + y1 + ' L ' + x2 + ' ' + y2)
        : ('M ' + x1 + ' ' + y1 + ' C ' + x1 + ' ' + midY + ', ' + x2 + ' ' + midY + ', ' + x2 + ' ' + y2);
      var path = document.createElementNS(svgNS, 'path');
      path.setAttribute('d', d);
      path.setAttribute('class', 'bz-edge');
      connectorsSvg.appendChild(path);
    });
  }

  // Während einer Auf-/Zuklapp-Animation die Verbindungen laufend neu zeichnen,
  // damit sie den animierten Zeilen exakt folgen.
  var connectorAnimStart = null;
  function animateConnectors() {
    var start = performance.now();
    connectorAnimStart = start;
    function step(ts) {
      if (connectorAnimStart !== start) return; // von einer neueren Animation überholt
      renderConnectors();
      if (ts - start < 260) {
        requestAnimationFrame(step);
      }
    }
    requestAnimationFrame(step);
  }

  renderConnectors();
  requestAnimationFrame(renderConnectors);
  setTimeout(renderConnectors, 80); // Sicherheitsnetz gegen späte Web-Font-Verschiebungen
  window.addEventListener('resize', renderConnectors);

  document.getElementById('bz-expand-all').addEventListener('click', function () {
    treeRoot.querySelectorAll('.bz-has-children').forEach(function (row) {
      row.classList.add('bz-open');
    });
    treeRoot.querySelectorAll('.bz-children').forEach(function (box) {
      setChildrenOpen(box, true);
    });
    animateConnectors();
  });
  document.getElementById('bz-collapse-all').addEventListener('click', function () {
    treeRoot.querySelectorAll('.bz-has-children').forEach(function (row) {
      row.classList.remove('bz-open');
    });
    treeRoot.querySelectorAll('.bz-children').forEach(function (box) {
      setChildrenOpen(box, false);
    });
    animateConnectors();
  });
})();
</script>
"""


st.title("CiteCut ✂️")
st.caption(
    "Wörter zählen nach formatierten Kapiteln – ohne Klammer-Zitate, ohne Tabellen."
)


# --------------------------------------------------------------------------- #
# 1. Upload
# --------------------------------------------------------------------------- #
uploaded = st.file_uploader(
    "Dokument hochladen (.docx oder .pdf)",
    type=["docx", "pdf"],
)

if uploaded is None:
    st.info("👆 Lade ein Word- oder PDF-Dokument hoch, um loszulegen.")
    st.stop()


@st.cache_data(show_spinner="Analysiere das Dokument …")
def _load(file_bytes: bytes, filename: str):
    """Parst das Dokument (gecached, damit Checkbox-Klicks schnell bleiben)."""
    root = parse_document(io.BytesIO(file_bytes), filename)
    # Nur die für die UI nötigen Titel der Hauptkapitel extrahieren.
    titles = [c.title for c in root.children]
    return root, titles


try:
    root, main_titles = _load(uploaded.getvalue(), uploaded.name)
except Exception as exc:  # noqa: BLE001
    st.error(f"Konnte das Dokument nicht lesen: {exc}")
    st.stop()

if not main_titles:
    st.warning(
        "Es wurden keine Kapitel/Überschriften gefunden. "
        "Bei Word: sind die Überschriften als Formatvorlage (Überschrift 1/2 …) formatiert? "
        "Bei PDF ist die Erkennung heuristisch."
    )
    st.stop()

if uploaded.name.lower().endswith(".pdf"):
    st.warning(
        "ℹ️ PDF-Erkennung ist heuristisch (Schriftgröße/Nummerierung). "
        "Für exakte Ergebnisse ist ein .docx mit echten Überschriften-Formatvorlagen zuverlässiger.",
        icon="⚠️",
    )


# --------------------------------------------------------------------------- #
# 2. Optionen (Checkboxen)
# --------------------------------------------------------------------------- #
st.subheader("Optionen")
col1, col2, col3 = st.columns(3)
with col1:
    include_headings = st.checkbox("Überschriften mitzählen", value=False)
with col2:
    include_parentheses = st.checkbox("Klammern mitzählen", value=False)
with col3:
    include_footnotes = st.checkbox("Fußnoten mitzählen", value=False)

opts = CountOptions(
    exclude_parentheses=not include_parentheses,
    include_headings=include_headings,
    include_footnotes=include_footnotes,
)


# --------------------------------------------------------------------------- #
# 3. Kapitel-Auswahl
# --------------------------------------------------------------------------- #
st.subheader("Kapitel-Auswahl")


def _chapter_label(number: str, title: str) -> str:
    """Kapitel genau mit der aus dem Dokument gelesenen Nummer anzeigen."""
    return f"{number} {title}".strip()


labels = [_chapter_label(ch.number, ch.title) for ch in root.children]
mode = st.radio(
    "Was soll gezählt werden?",
    ["Alle Kapitel", "Bereich (von–bis)", "Einzelauswahl"],
    horizontal=True,
)

selected_indices = list(range(len(main_titles)))

if mode == "Bereich (von–bis)":
    c1, c2 = st.columns(2)
    with c1:
        i_start = st.selectbox(
            "Von Kapitel",
            range(len(labels)),
            index=0,
            format_func=lambda index: labels[index],
        )
    with c2:
        i_end = st.selectbox(
            "Bis Kapitel",
            range(len(labels)),
            index=len(labels) - 1,
            format_func=lambda index: labels[index],
        )
    if i_start > i_end:
        i_start, i_end = i_end, i_start
    selected_indices = list(range(i_start, i_end + 1))
elif mode == "Einzelauswahl":
    selected_indices = st.multiselect(
        "Kapitel wählen",
        range(len(labels)),
        default=list(range(len(labels))),
        format_func=lambda index: labels[index],
    )

if not selected_indices:
    st.warning("Bitte mindestens ein Kapitel auswählen.")
    st.stop()


# --------------------------------------------------------------------------- #
# 4. Zählen & Ausgabe
# --------------------------------------------------------------------------- #
with st.spinner("✂️ Zähle die Kapitel …"):
    nodes = count_document(root, opts, selected_indices)
    total = grand_total(nodes)
    time.sleep(0.45)  # kurze Verzögerung, damit der Spinner sichtbar bleibt

st.divider()
st.metric("Gesamtwortzahl (Auswahl)", f"{total:,}".replace(",", "."))


def _pct(part: int, whole: int) -> float:
    return round((part / whole * 100), 1) if whole else 0.0


def _node_to_dict(node: CountNode, main_total: int, selection_total: int, is_main: bool) -> dict:
    """Baumknoten als JSON-taugliches dict, inkl. vorberechnetem Prozentanteil.

    Hauptkapitel (``is_main``): Anteil an ``selection_total`` (der Auswahl).
    Alle Unterkapitel, egal auf welcher Tiefe: Anteil am eigenen Hauptkapitel
    (``main_total``), nicht am jeweils direkt übergeordneten Abschnitt.
    """
    pct = _pct(node.total, selection_total if is_main else main_total)
    effective_main_total = node.total if is_main else main_total
    return {
        "number": node.number,
        "title": node.title,
        "own": node.own,
        "total": node.total,
        "pct": pct,
        "children": [
            _node_to_dict(child, effective_main_total, selection_total, False)
            for child in node.children
        ],
    }


def _count_rows(node: dict, is_main: bool) -> int:
    """Zählt sichtbare Zeilen im voll aufgeklappten Baum (für die Komponentenhöhe)."""
    count = 1
    if is_main and node["children"] and node["own"] > 0:
        count += 1  # "eigener Text"-Zeile
    for child in node["children"]:
        count += _count_rows(child, False)
    return count


st.subheader("Kapitel im Detail")

tree_data = [_node_to_dict(node, 0, total, True) for node in nodes]
tree_json = json.dumps(tree_data, ensure_ascii=False).replace("</", "<\\/")

total_rows = sum(_count_rows(node, True) for node in tree_data)
# Kein Scroll-Fenster: die Komponente wird so hoch wie der voll aufgeklappte Baum.
BAR_HEIGHT_PX = 100
component_height = BAR_HEIGHT_PX + 44 + 22 + total_rows * 36 + 24
tree_html = _TREE_HTML.replace("__DATA__", tree_json).replace("__BAR_HEIGHT__", str(BAR_HEIGHT_PX))
components.html(tree_html, height=component_height, scrolling=False)


# --------------------------------------------------------------------------- #
# 5. Export
# --------------------------------------------------------------------------- #
st.divider()
st.subheader("Export")

_opts_label = (
    f"Überschriften={'ja' if include_headings else 'nein'}, "
    f"Klammern={'ja' if include_parentheses else 'nein'}, "
    f"Fußnoten={'ja' if include_footnotes else 'nein'}"
)


def _flatten(nodes):
    """Kapitelbaum in flache Zeilen umwandeln (für CSV/Excel)."""
    rows = []

    def walk(node: CountNode, is_main: bool):
        rows.append(
            {
                "Nummer": node.number,
                "Ebene": node.level,
                "Hauptkapitel": "ja" if is_main else "nein",
                "Kapitel": node.title,
                "Woerter_direkt": node.own,
                "Woerter_gesamt": node.total,
            }
        )
        for child in node.children:
            walk(child, False)

    for n in nodes:
        walk(n, True)
    return rows


rows = _flatten(nodes)
columns = ["Nummer", "Ebene", "Hauptkapitel", "Kapitel", "Woerter_direkt", "Woerter_gesamt"]


# ---- TXT ---------------------------------------------------------------- #
def _report_lines(node: CountNode, depth: int = 0):
    indent = "    " * depth
    heading = _chapter_label(node.number, node.title)
    lines = [f"{indent}{heading}: {node.total} Wörter"]
    if node.children and node.own > 0:
        lines.append(f"{indent}    Fließtext direkt: {node.own} Wörter")
    for child in node.children:
        lines.extend(_report_lines(child, depth + 1))
    return lines


report = [
    "CiteCut – Auswertung",
    f"Datei: {uploaded.name}",
    f"Optionen: {_opts_label}",
    "",
    f"GESAMTWORTZAHL (Auswahl): {total}",
    "",
]
for node in nodes:
    report.extend(_report_lines(node))
    report.append("")
txt_bytes = "\n".join(report).encode("utf-8")


# ---- CSV (Semikolon + BOM, damit Excel Umlaute richtig zeigt) ----------- #
def _build_csv() -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=";", lineterminator="\r\n")
    writer.writerow([f"CiteCut – {uploaded.name}"])
    writer.writerow([f"Optionen: {_opts_label}"])
    writer.writerow([])
    writer.writerow(columns)
    for r in rows:
        writer.writerow([r[c] for c in columns])
    writer.writerow([])
    writer.writerow(["", "", "", "GESAMT (Auswahl)", "", total])
    return ("﻿" + buf.getvalue()).encode("utf-8")


# ---- Excel -------------------------------------------------------------- #
def _build_xlsx() -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Wortzählung"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F6FEB")
    main_font = Font(bold=True)

    ws.append(["CiteCut-Auswertung"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([f"Datei: {uploaded.name}"])
    ws.append([f"Optionen: {_opts_label}"])
    ws.append([])

    header_row = ws.max_row + 1
    ws.append(columns)
    for cell in ws[header_row]:
        cell.font = header_font
        cell.fill = header_fill

    for r in rows:
        ws.append([r[c] for c in columns])
        if r["Hauptkapitel"] == "ja":
            for cell in ws[ws.max_row]:
                cell.font = main_font
        else:
            # Unterkapitel im Titel einrücken.
            ws.cell(row=ws.max_row, column=4).alignment = Alignment(indent=r["Ebene"] - 1)

    ws.append([])
    ws.append(["", "", "", "GESAMT (Auswahl)", "", total])
    for cell in ws[ws.max_row]:
        cell.font = main_font

    widths = {"A": 8, "B": 7, "C": 13, "D": 42, "E": 15, "F": 15}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


c1, c2, c3 = st.columns(3)
with c1:
    st.download_button(
        "📄 .txt",
        data=txt_bytes,
        file_name="citecut_report.txt",
        mime="text/plain",
        use_container_width=True,
    )
with c2:
    st.download_button(
        "📊 .csv",
        data=_build_csv(),
        file_name="citecut.csv",
        mime="text/csv",
        use_container_width=True,
    )
with c3:
    st.download_button(
        "📗 .xlsx",
        data=_build_xlsx(),
        file_name="citecut.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
